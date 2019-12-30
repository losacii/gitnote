from openpyxl import load_workbook
wb = load_workbook("./hello.xlsx")
ws = wb[wb.sheetnames[0]]

ws['G5'] = "G5"

for i in range(8):
    pos = 'C{}'.format(i+1,)
    ws[pos] = (i + 1)

    pos = 'D{}'.format(i+1,)
    ws[pos] = (i + 1) ** 2

ws2 = wb.create_sheet("foo1")
ws3 = wb.create_sheet('foo2')
ws4 = wb.create_sheet('foo3')
ws3 = wb.create_sheet('bar1', -1)
ws3 = wb.create_sheet('bar2', -1)
ws3 = wb.create_sheet('bar3', -1)

ws5 = wb.copy_worksheet(ws)
ws5['G6'] = 'G6'

wb.save("new.xlsx")

# ws = wb.active

# ws["A1"] = "hello"
# ws["B1"] = "world!"

# 
